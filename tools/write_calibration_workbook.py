import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


OUTPUT_DIR = Path("outputs/scoring-audit")
JSON_PATH = OUTPUT_DIR / "hermancoach_scoring_calibration_loops.json"
XLSX_PATH = OUTPUT_DIR / "hermancoach_scoring_calibration_loops.xlsx"


def add_table(ws, headers, rows):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F2A44")
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append(row)
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        width = min(max(max_length + 2, 10), 54)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def main():
    payload = json.loads(JSON_PATH.read_text())
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    combined = payload["combinedStats"]
    add_table(
        summary,
        ["Metric", "Value"],
        [
            ["Total prompts", combined["count"]],
            ["Average miss %", combined["avgMiss"]],
            ["Max miss %", combined["maxMiss"]],
            ["Min miss %", combined["minMiss"]],
            ["Std dev miss %", combined["stdDev"]],
            ["Average absolute delta", combined["avgAbsDelta"]],
            ["Review cases", combined["reviewCount"]],
            ["Stop condition met", "Yes" if combined["avgMiss"] < 8 or combined["maxMiss"] < 25 else "No"],
        ],
    )

    pass_stats = wb.create_sheet("Pass Stats")
    add_table(
        pass_stats,
        [
            "Pass",
            "Count",
            "Average Miss %",
            "Max Miss %",
            "Min Miss %",
            "Std Dev Miss %",
            "Average Abs Delta",
            "Review Cases",
        ],
        [
            [
                row["pass"],
                row["count"],
                row["avgMiss"],
                row["maxMiss"],
                row["minMiss"],
                row["stdDev"],
                row["avgAbsDelta"],
                row["reviewCount"],
            ]
            for row in payload["passStats"]
        ],
    )

    high_misses = wb.create_sheet("Highest Misses")
    add_table(
        high_misses,
        ["ID", "Pass", "Task Type", "Quality", "Heuristic Score", "Expected Score", "Miss %", "Prompt"],
        [
            [
                row["id"],
                row["pass"],
                row["type"],
                row["quality"],
                row["heuristic"],
                row["expected"],
                row["miss"],
                row["prompt"],
            ]
            for row in payload["highMisses"]
        ],
    )

    audit = wb.create_sheet("Audit Results")
    add_table(
        audit,
        [
            "ID",
            "Pass",
            "Prompt Snippet",
            "Prompt Description",
            "Task Type",
            "Quality",
            "Heuristic Score",
            "Expected Score",
            "Delta",
            "Abs Delta",
            "Miss %",
            "Miss Status",
            "Who",
            "Task",
            "Context",
            "Output",
            "Root Cause",
            "Specific Rule Update",
            "Full Prompt",
        ],
        [
            [
                row["id"],
                row["pass"],
                row["prompt"][:110],
                row["description"],
                row["type"],
                row["quality"],
                row["heuristicScore"],
                row["expectedScore"],
                row["delta"],
                row["absDelta"],
                row["miss"],
                row["status"],
                row["who"],
                row["task"],
                row["context"],
                row["output"],
                row["rootCause"],
                row["specificRule"],
                row["prompt"],
            ]
            for row in payload["rows"]
        ],
    )

    wb.save(XLSX_PATH)
    print(XLSX_PATH)


if __name__ == "__main__":
    main()
