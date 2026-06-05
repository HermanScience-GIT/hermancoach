import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


OUTPUT_DIR = Path("outputs/scoring-audit")
JSON_PATH = OUTPUT_DIR / "hermancoach_scoring_confirmation_holdout.json"
XLSX_PATH = OUTPUT_DIR / "hermancoach_scoring_confirmation_holdout.xlsx"


def add_table(ws, headers, rows):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F2A44")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in rows:
        ws.append(row)
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        width = min(max(max_length + 2, 10), 58)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def main():
    payload = json.loads(JSON_PATH.read_text())
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    combined = payload["combinedStats"]
    add_table(
        summary,
        ["Metric", "Value"],
        [
            ["Total prompts", combined["count"]],
            ["Average miss %", combined["avgMiss"]],
            ["Max miss %", combined["maxMiss"]],
            ["Min miss %", combined["minMiss"]],
            ["Std dev miss %", combined["stdDev"]],
            ["Average absolute delta", combined["avgAbsDelta"]],
            ["Review cases", combined["reviewCount"]],
            ["Previous stop condition confirmed", "Yes" if combined["avgMiss"] < 8 or combined["maxMiss"] < 25 else "No"],
        ],
    )

    high_misses = wb.create_sheet("Highest Misses")
    add_table(
        high_misses,
        [
            "ID",
            "Task Type",
            "Description",
            "Heuristic Score",
            "Expected Score",
            "Delta",
            "Miss %",
            "Root Cause",
            "Specific Rule Update",
            "Prompt",
        ],
        [
            [
                row["id"],
                row["type"],
                row["description"],
                row["heuristicScore"],
                row["expectedScore"],
                row["delta"],
                row["miss"],
                row["rootCause"],
                row["specificRule"],
                row["prompt"],
            ]
            for row in payload["highMisses"]
        ],
    )

    audit = wb.create_sheet("Audit Results")
    add_table(
        audit,
        [
            "ID",
            "Prompt Snippet",
            "Prompt Description",
            "Task Type",
            "Heuristic Score",
            "Expected Score",
            "Delta",
            "Abs Delta",
            "Miss %",
            "Miss Status",
            "Who",
            "Task",
            "Context",
            "Output",
            "Root Cause",
            "Specific Rule Update",
            "Full Prompt",
        ],
        [
            [
                row["id"],
                row["prompt"][:110],
                row["description"],
                row["type"],
                row["heuristicScore"],
                row["expectedScore"],
                row["delta"],
                row["absDelta"],
                row["miss"],
                row["status"],
                row["who"],
                row["task"],
                row["context"],
                row["output"],
                row["rootCause"],
                row["specificRule"],
                row["prompt"],
            ]
            for row in payload["rows"]
        ],
    )

    wb.save(XLSX_PATH)
    print(XLSX_PATH)


if __name__ == "__main__":
    main()
